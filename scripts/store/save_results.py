import json
from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter


def _compute_term_quarters(raw_facts: dict, tags: list, prev_end: str, this_end: str):
    """
    For a given set of GAAP tags, compute Q1–Q3 (min of all 10-Qs between prev_end and this_end)
    and the FY total (10-K at this_end). Returns (q_vals, total_val, q4_val).
    """
    # Find FY-end total
    total_val = 0
    for tag in tags:
        fact = raw_facts.get(tag) or raw_facts.get(tag.split(':', 1)[-1])
        if not fact:
            continue
        for e in fact.get('units', {}).get('USD', []):
            if e.get('form') == '10-K' and e.get('end') == this_end:
                total_val = e.get('val') or 0

    # Bucket Q1–Q3
    buckets = {'Q1': [], 'Q2': [], 'Q3': []}
    for tag in tags:
        fact = raw_facts.get(tag) or raw_facts.get(tag.split(':', 1)[-1])
        if not fact:
            continue
        for e in fact.get('units', {}).get('USD', []):
            fp = e.get('fp')
            if e.get('form') == '10-Q' and fp in buckets:
                end = e.get('end')
                val = e.get('val')
                if end and val is not None and prev_end < end < this_end:
                    buckets[fp].append(val)

    q_vals = {q: (min(vals) if vals else 0) for q, vals in buckets.items()}
    q4 = total_val - sum(q_vals.values())
    return q_vals, total_val, q4


def save_results(df_matched: pd.DataFrame,
                 mapping_path: str,
                 out_path: str,
                 fy_map_override: dict = None) -> None:
    """
    Writes three sheets—Income, Balance, Cashflow—with FY columns side-by-side.
    Each FY produces columns: <FY>-10K, <FY>-Q1, <FY>-Q2, <FY>-Q3, <FY>-Q4 (values in millions).
    Fiscal years are in descending order (latest first).
    """
    # Infer CIK from output filename
    cik = Path(out_path).stem.split('_')[0]

    # Load mapping and raw facts
    mapping = json.load(Path(mapping_path).open())
    raw = json.loads(Path(f"data/raw/{cik}.json").read_text())
    raw_facts = raw.get('facts', {}).get('us-gaap', {})
    
    # 1) Gather all 10-K records (fy, end) and map to latest end per FY
    # ─── override if provided ─────────────────────────────────────────────
    if fy_map_override is not None:
        fy_map = fy_map_override
    else:
        # 1) Gather all 10-K records
        k_recs = []
        for section in mapping.values():
            for tags in section.values():
                for tag in tags:
                    fact = raw_facts.get(tag) or raw_facts.get(tag.split(":",1)[-1])
                    if not fact:
                        continue
                    for e in fact.get("units", {}).get("USD", []):
                        if e.get("form") == "10-K" and e.get("fy") and e.get("end"):
                            k_recs.append((int(e["fy"]), e["end"], e.get("val") or 0))

        # 2) Build fy_map from those records
        fy_map = {}
        for fy, end, _ in k_recs:
            if fy not in fy_map or end > fy_map[fy]:
                fy_map[fy] = end

    # Descending years: latest first
    years = sorted(fy_map.keys(), reverse=True)
    if not years:
        raise ValueError(f"No fiscal years found for {cik}")

    # Prepare the column order: for each FY (desc), 10-K, Q1, Q2, Q3, Q4
    all_cols = []
    for fy in years:
        all_cols += [f"{fy}-10K", f"{fy}-Q1", f"{fy}-Q2", f"{fy}-Q3", f"{fy}-Q4"]

    # 2) Build each statement sheet
    sheets = []
    for sheet_name, section_key in [
        ("Income Statement",   "income_statement"),
        ("Balance Sheet",      "balance_sheet"),
        ("Cashflow Statement", "cashflow_statement")
    ]:
        rows = []
        for std_term, tags in mapping.get(section_key, {}).items():
            rec = {"standard_term": std_term}
            for fy in years:
                prev_end = fy_map.get(fy - 1)
                this_end = fy_map[fy]
                if prev_end is None:
                    # no prior FY: all zeros
                    rec[f"{fy}-10K"] = 0
                    rec[f"{fy}-Q1"]  = 0
                    rec[f"{fy}-Q2"]  = 0
                    rec[f"{fy}-Q3"]  = 0
                    rec[f"{fy}-Q4"]  = 0
                else:
                    q_vals, total, q4 = _compute_term_quarters(raw_facts, tags, prev_end, this_end)
                    rec[f"{fy}-10K"] = total or 0
                    rec[f"{fy}-Q1"]  = q_vals.get("Q1", 0)
                    rec[f"{fy}-Q2"]  = q_vals.get("Q2", 0)
                    rec[f"{fy}-Q3"]  = q_vals.get("Q3", 0)
                    rec[f"{fy}-Q4"]  = q4
            rows.append(rec)

        df_sheet = pd.DataFrame(rows).set_index("standard_term") / 1e6
        df_sheet = df_sheet[all_cols]
        sheets.append((sheet_name, df_sheet))

    # 3) Write to Excel with three sheets
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df_sheet in sheets:
            df_sheet.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=2,
                startcol=0,
                header=False
            )
            ws = writer.sheets[sheet_name]

            # Row 1: FY labels (desc order)
            for idx, col in enumerate(df_sheet.columns, start=2):
                fy_label = col.split("-", 1)[0]
                ws.cell(row=1, column=idx, value=f"FY {fy_label}")

            # Row 2: sub-period labels
            for idx, col in enumerate(df_sheet.columns, start=2):
                sub = col.split("-", 1)[1]
                ws.cell(row=2, column=idx, value=sub)

            # Auto-adjust column widths
            total_rows = 3 + len(df_sheet)
            total_cols = 1 + len(df_sheet.columns)
            for col_idx in range(1, total_cols + 1):
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, total_rows + 1)
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    print(f"Results saved to {out_path}")
        
    